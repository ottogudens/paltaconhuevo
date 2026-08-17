import logging

from django.conf import settings
from django.contrib.auth import get_user_model
from django.db.models import Q
from django.http import HttpResponse
import openpyxl
from rest_framework import generics, status
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework.response import Response
from rest_framework.authtoken.models import Token
from rest_framework.views import APIView

from core.permissions import IsAdmin, IsAdminOrVendedor, IsWhatsAppService
from .models import PasswordResetToken
from .serializers import (
    UserSerializer,
    RegisterSerializer,
    LoginSerializer,
    CreateUserSerializer,
)

User = get_user_model()
logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _phone_variants(phone: str) -> list:
    """Genera variantes normalizadas de un número de teléfono chileno."""
    clean = ''.join(filter(str.isdigit, phone))
    variants = [phone]
    if clean:
        variants.append(clean)
        if len(clean) == 9:
            variants += [f"+56{clean}", f"56{clean}"]
        elif len(clean) == 11 and clean.startswith('569'):
            variants += [clean[2:], f"+{clean}"]
    return list(dict.fromkeys(variants))  # deduplica manteniendo orden


def _send_reset_email(user, token_str: str) -> bool:
    """Envía el email de recuperación con SendGrid. Retorna True si tuvo éxito."""
    try:
        import sendgrid
        from sendgrid.helpers.mail import Mail

        api_key = getattr(settings, 'SENDGRID_API_KEY', '')
        from_email = getattr(settings, 'SENDGRID_FROM_EMAIL', 'noreply@paltaconhuevo.cl')
        if not api_key:
            logger.error("SENDGRID_API_KEY no configurada; no se pudo enviar email de reset.")
            return False

        message = Mail(
            from_email=from_email,
            to_emails=user.email,
            subject="Recuperación de contraseña — Palta con Huevo",
            html_content=(
                f"<p>Hola {user.first_name or user.username},</p>"
                f"<p>Tu código de recuperación de contraseña es:</p>"
                f"<h2 style='letter-spacing:4px'>{token_str}</h2>"
                f"<p>Este código expira en <strong>15 minutos</strong>.</p>"
                f"<p>Si no solicitaste este cambio, ignora este mensaje.</p>"
                f"<p>— Equipo Palta con Huevo 🥑</p>"
            ),
        )
        sg = sendgrid.SendGridAPIClient(api_key=api_key)
        sg.send(message)
        return True
    except Exception as exc:
        logger.exception("Error enviando email de reset: %s", exc)
        return False


# ---------------------------------------------------------------------------
# Auth público
# ---------------------------------------------------------------------------

class RegisterView(generics.CreateAPIView):
    serializer_class = RegisterSerializer
    permission_classes = [AllowAny]

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        user = serializer.save()
        token, _ = Token.objects.get_or_create(user=user)
        return Response(
            {'token': token.key, 'user': UserSerializer(user).data},
            status=status.HTTP_201_CREATED,
        )


class LoginView(APIView):
    permission_classes = [AllowAny]

    def post(self, request):
        serializer = LoginSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        user = serializer.validated_data['user']
        token, _ = Token.objects.get_or_create(user=user)
        return Response({'token': token.key, 'user': UserSerializer(user).data})


# ---------------------------------------------------------------------------
# C2 — WhatsAppAuthView: solo el agente Node puede llamarlo
# ---------------------------------------------------------------------------

class WhatsAppAuthView(APIView):
    """
    Endpoint exclusivo para el agente WhatsApp.
    Protegido con IsAdmin: solo el agente Node (autenticado con el token DRF
    de un usuario admin — DJANGO_API_TOKEN) puede llamar este endpoint.
    Los clientes normales no pueden obtener tokens de otros usuarios.
    """
    permission_classes = [IsAdmin]

    def post(self, request):
        phone = str(request.data.get('phone', '')).strip()
        name = str(request.data.get('name', '')).strip()

        if not phone:
            return Response(
                {'error': 'El teléfono es requerido'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        phone_variants = _phone_variants(phone)
        clean_digits = ''.join(filter(str.isdigit, phone))

        user = User.objects.filter(
            Q(phone__in=phone_variants)
            | Q(whatsapp_number__in=phone_variants)
            | Q(username__in=phone_variants)
        ).first()

        if user:
            if name and (
                not user.first_name
                or user.first_name.startswith('569')
                or user.first_name == 'wa'
            ):
                parts = name.split(' ')
                user.first_name = parts[0]
                user.last_name = ' '.join(parts[1:]) if len(parts) > 1 else ''
                user.save(update_fields=['first_name', 'last_name'])
            token, _ = Token.objects.get_or_create(user=user)
            return Response({
                'token': token.key,
                'user': UserSerializer(user).data,
                'is_new': False,
            })

        if not name:
            return Response(
                {'name_required': True, 'message': 'Se requiere el nombre del usuario para el registro'},
                status=status.HTTP_200_OK,
            )

        parts = name.split(' ')
        first_name = parts[0]
        last_name = ' '.join(parts[1:]) if len(parts) > 1 else ''
        username = f"wa_{clean_digits or phone}"

        if User.objects.filter(username=username).exists():
            import uuid
            username = f"{username}_{uuid.uuid4().hex[:4]}"

        user = User.objects.create_user(
            username=username,
            email=f"{clean_digits or phone}@whatsapp.cl",
            first_name=first_name,
            last_name=last_name,
            phone=phone,
            whatsapp_number=phone,
            role='cliente',
        )
        # C2 fix: contraseña aleatoria, no derivada del número de teléfono
        user.set_password(User.objects.make_random_password(length=24))
        user.save()

        from loyalty.models import LoyaltyAccount
        LoyaltyAccount.objects.get_or_create(user=user)

        token, _ = Token.objects.get_or_create(user=user)
        return Response(
            {'token': token.key, 'user': UserSerializer(user).data, 'is_new': True},
            status=status.HTTP_201_CREATED,
        )


# ---------------------------------------------------------------------------
# Logout / Perfil
# ---------------------------------------------------------------------------

class LogoutView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        request.user.auth_token.delete()
        return Response({'message': 'Sesión cerrada'})


class ProfileView(generics.RetrieveUpdateAPIView):
    serializer_class = UserSerializer
    permission_classes = [IsAuthenticated]

    def get_object(self):
        return self.request.user


# ---------------------------------------------------------------------------
# C3 — Reset de contraseña en dos pasos con token temporal
# ---------------------------------------------------------------------------

class PasswordResetRequestView(APIView):
    """
    Paso 1: el usuario envía su email o teléfono.
    Genera un token y lo envía por email.
    Siempre responde con el mismo mensaje genérico para evitar
    enumeración de usuarios.
    """
    permission_classes = [AllowAny]

    _GENERIC_OK = {'message': 'Si existe una cuenta asociada recibirás un email con instrucciones.'}

    def post(self, request):
        identifier = (request.data.get('identifier') or '').strip()
        if not identifier:
            return Response({'error': 'Debes proporcionar tu correo o teléfono.'}, status=400)

        phone_variants = _phone_variants(identifier)
        user = User.objects.filter(
            Q(email__iexact=identifier)
            | Q(username__iexact=identifier)
            | Q(phone__in=phone_variants)
            | Q(whatsapp_number__in=phone_variants)
        ).first()

        if not user:
            # Respuesta genérica — no confirmar si el usuario existe
            return Response(self._GENERIC_OK)

        if not user.email or '@whatsapp.cl' in user.email:
            # Clientes creados por WhatsApp sin email real no pueden recibir reset
            return Response(self._GENERIC_OK)

        reset_token = PasswordResetToken.generate_for(user)
        sent = _send_reset_email(user, reset_token.token)
        if not sent:
            logger.error("No se pudo enviar email de reset a user_id=%s", user.id)

        return Response(self._GENERIC_OK)


class PasswordResetConfirmView(APIView):
    """
    Paso 2: el usuario envía el token recibido y su nueva contraseña.
    El token expira en 15 minutos y es de un solo uso.
    """
    permission_classes = [AllowAny]

    def post(self, request):
        token_str = (request.data.get('token') or '').strip()
        new_password = (request.data.get('new_password') or '').strip()

        if not token_str or not new_password:
            return Response({'error': 'Se requiere el token y la nueva contraseña.'}, status=400)

        if len(new_password) < 8:
            return Response({'error': 'La contraseña debe tener al menos 8 caracteres.'}, status=400)

        try:
            reset_token = PasswordResetToken.objects.select_related('user').get(token=token_str)
        except PasswordResetToken.DoesNotExist:
            return Response({'error': 'Token inválido o expirado.'}, status=400)

        if not reset_token.is_valid:
            return Response({'error': 'Token inválido o expirado.'}, status=400)

        reset_token.user.set_password(new_password)
        reset_token.user.save(update_fields=['password'])
        reset_token.used = True
        reset_token.save(update_fields=['used'])

        return Response({'message': 'Contraseña restablecida con éxito. Ya puedes iniciar sesión.'})


# ---------------------------------------------------------------------------
# Gestión de clientes — solo staff (C1)
# ---------------------------------------------------------------------------

class CustomerListView(generics.ListCreateAPIView):
    permission_classes = [IsAdminOrVendedor]

    def get_serializer_class(self):
        if self.request.method == 'POST':
            return CreateUserSerializer
        return UserSerializer

    def get_queryset(self):
        return User.objects.filter(role='cliente').order_by('-created_at')

    def perform_create(self, serializer):
        serializer.save(role='cliente')


class CustomerDetailView(generics.RetrieveUpdateDestroyAPIView):
    permission_classes = [IsAdminOrVendedor]
    queryset = User.objects.filter(role='cliente')

    def get_serializer_class(self):
        if self.request.method in ('PUT', 'PATCH'):
            return CreateUserSerializer
        return UserSerializer


# ---------------------------------------------------------------------------
# Gestión de usuarios del sistema — solo admin (C1)
# ---------------------------------------------------------------------------

class SystemUserListView(generics.ListCreateAPIView):
    permission_classes = [IsAdmin]

    def get_serializer_class(self):
        if self.request.method == 'POST':
            return CreateUserSerializer
        return UserSerializer

    def get_queryset(self):
        qs = User.objects.filter(role__in=['admin', 'vendedor']).order_by('-created_at')
        role = self.request.query_params.get('role')
        if role and role in ('admin', 'vendedor'):
            qs = qs.filter(role=role)
        return qs


class SystemUserDetailView(generics.RetrieveUpdateDestroyAPIView):
    permission_classes = [IsAdmin]
    queryset = User.objects.filter(role__in=['admin', 'vendedor'])

    def get_serializer_class(self):
        if self.request.method in ('PUT', 'PATCH'):
            return CreateUserSerializer
        return UserSerializer


# ---------------------------------------------------------------------------
# Exportación / Importación de clientes — staff (C1)
# ---------------------------------------------------------------------------

class ExportCustomersView(APIView):
    permission_classes = [IsAdminOrVendedor]

    def get(self, request):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Clientes"
        headers = [
            'ID', 'Nombre', 'Email', 'Teléfono', 'WhatsApp',
            'Dirección', 'Comuna', 'Método de pago', 'Condición pago', 'Fecha registro',
        ]
        ws.append(headers)
        for u in User.objects.filter(role='cliente'):
            ws.append([
                u.id, u.get_full_name(), u.email, u.phone,
                u.whatsapp_number, u.address, u.commune,
                u.preferred_payment_method, u.preferred_payment_condition,
                str(u.created_at.date()),
            ])
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="clientes.xlsx"'
        wb.save(response)
        return response


class DownloadCustomerTemplateView(APIView):
    permission_classes = [IsAdminOrVendedor]

    def get(self, request):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Plantilla Clientes"
        headers = [
            'ID', 'Nombre', 'Email', 'Teléfono', 'WhatsApp',
            'Dirección', 'Comuna', 'Método de pago', 'Condición pago', 'Fecha registro'
        ]
        ws.append(headers)
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="plantilla_clientes.xlsx"'
        wb.save(response)
        return response


class ImportCustomersView(APIView):
    permission_classes = [IsAdmin]

    def post(self, request):
        file = request.FILES.get('file')
        if not file:
            return Response({'error': 'No se recibió archivo'}, status=400)
        wb = openpyxl.load_workbook(file)
        ws = wb.active
        created = 0
        errors = []
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                # Índices mapeados según la exportación:
                # 0: ID, 1: Nombre, 2: Email, 3: Teléfono
                name_parts = (str(row[1] or '')).split(' ', 1)
                email = str(row[2] or '').strip()
                phone = str(row[3] or '').strip() if len(row) > 3 else ''
                whatsapp = str(row[4] or '').strip() if len(row) > 4 else ''
                address = str(row[5] or '').strip() if len(row) > 5 else ''
                commune = str(row[6] or '').strip() if len(row) > 6 else ''

                if not name_parts[0]:
                    errors.append(f"Fila {i}: Falta el Nombre")
                    continue
                if not whatsapp:
                    errors.append(f"Fila {i}: Falta el WhatsApp")
                    continue

                if not email:
                    clean_whatsapp = ''.join(filter(str.isdigit, whatsapp))
                    email = f"{clean_whatsapp}@whatsapp.cl"

                defaults_data = {
                    'username': email.split('@')[0],
                    'first_name': name_parts[0],
                    'last_name': name_parts[1] if len(name_parts) > 1 else '',
                    'phone': phone,
                    'whatsapp_number': whatsapp,
                    'address': address,
                    'commune': commune,
                    'role': 'cliente',
                }

                user, c = User.objects.get_or_create(
                    email=email,
                    defaults=defaults_data,
                )
                if c:
                    # Contraseña aleatoria — no la misma para todos (A6 fix parcial)
                    user.set_password(User.objects.make_random_password(length=20))
                    user.save()
                    from loyalty.models import LoyaltyAccount
                    LoyaltyAccount.objects.get_or_create(user=user)
                    created += 1
                else:
                    # Actualizar usuario existente
                    updated = False
                    for field, value in defaults_data.items():
                        if field != 'username' and getattr(user, field) != value:
                            setattr(user, field, value)
                            updated = True
                    if updated:
                        user.save()
            except Exception as e:
                errors.append(f"Fila {i}: {str(e)}")
        return Response({'created': created, 'errors': errors})
