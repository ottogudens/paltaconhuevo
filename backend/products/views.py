from rest_framework import generics, status
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated, IsAuthenticatedOrReadOnly
from django.http import HttpResponse
import openpyxl
from .models import Product, Purchase
from .serializers import ProductSerializer, PurchaseSerializer
from core.permissions import IsAdminOrVendedor


class ProductListCreateView(generics.ListCreateAPIView):
    """
    GET: lectura pública para clientes autenticados y el agente WhatsApp.
    POST (crear producto): solo staff.
    """
    serializer_class = ProductSerializer
    queryset = Product.objects.filter(is_active=True).order_by('name')

    def get_queryset(self):
        qs = Product.objects.filter(is_active=True).order_by('name')
        user = self.request.user
        if hasattr(user, 'role') and user.role in ['admin', 'vendedor']:
            return qs
        return qs.filter(can_be_sold=True)

    def get_permissions(self):
        if self.request.method == 'POST':
            return [IsAdminOrVendedor()]
        return [IsAuthenticated()]

    def create(self, request, *args, **kwargs):
        return super().create(request, *args, **kwargs)


class ProductDetailView(generics.RetrieveUpdateDestroyAPIView):
    """Modificar/eliminar productos: solo staff."""
    serializer_class = ProductSerializer
    queryset = Product.objects.all()
    permission_classes = [IsAdminOrVendedor]

    def update(self, request, *args, **kwargs):
        return super().update(request, *args, **kwargs)


class PurchaseListCreateView(generics.ListCreateAPIView):
    serializer_class = PurchaseSerializer
    queryset = Purchase.objects.all().order_by('-purchase_date')
    permission_classes = [IsAdminOrVendedor]


class PurchaseDetailView(generics.RetrieveUpdateDestroyAPIView):
    serializer_class = PurchaseSerializer
    queryset = Purchase.objects.all()
    permission_classes = [IsAdminOrVendedor]


class ExportPurchasesView(APIView):
    permission_classes = [IsAdminOrVendedor]

    def get(self, request):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Compras"
        ws.append(['ID', 'Proveedor', 'Producto', 'Cantidad', 'Costo unitario', 'Costo total', 'Fecha'])
        for p in Purchase.objects.all().order_by('-purchase_date'):
            ws.append([p.id, p.supplier_name, p.product.name, float(p.quantity),
                       float(p.unit_cost), float(p.total_cost), str(p.purchase_date)])
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="compras.xlsx"'
        wb.save(response)
        return response


class LowStockView(APIView):
    permission_classes = [IsAdminOrVendedor]

    def get(self, request):
        products = Product.objects.filter(is_active=True)
        low = [ProductSerializer(p).data for p in products if p.stock <= p.min_stock]
        return Response(low)

class DownloadProductTemplateView(APIView):
    permission_classes = [IsAdminOrVendedor]

    def get(self, request):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Plantilla Productos"
        # IMPORTANTE: el orden de estas columnas debe coincidir exactamente
        # con lo que lee ImportProductsView (min_row=2, por índice de columna).
        headers = [
            'Nombre',           # col 0 — obligatorio
            'Descripción',      # col 1
            'Tipo',             # col 2 — palta | huevo | otro
            'Precio Venta',     # col 3
            'Precio Compra',    # col 4
            'Stock Actual',     # col 5
            'Stock Mínimo',     # col 6
            'Unidad de Medida', # col 7 — unidad | kilo | docena | caja
            'Activo',           # col 8 — true | false
        ]
        ws.append(headers)

        # Fila de ejemplo para guiar al usuario
        ws.append([
            'Palta Hass', 'Palta de primera calidad', 'palta',
            2500, 1200, 50, 10, 'kilo', 'true'
        ])

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="plantilla_productos.xlsx"'
        wb.save(response)
        return response


class ImportProductsView(APIView):
    permission_classes = [IsAdminOrVendedor]

    def post(self, request):
        file = request.FILES.get('file')
        import_mode = request.data.get('import_mode', 'update')

        if not file:
            return Response({'error': 'No se recibió archivo'}, status=400)

        if import_mode == 'replace':
            Product.objects.filter(is_bundle=False).delete()

        wb = openpyxl.load_workbook(file)
        ws = wb.active
        created = 0
        updated = 0
        errors = []

        VALID_TYPES = {'palta', 'huevo', 'otro'}
        VALID_UNITS = {'unidad', 'kilo', 'docena', 'caja'}

        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                # Columnas según la plantilla (mismo orden):
                # 0: Nombre, 1: Descripción, 2: Tipo, 3: Precio Venta,
                # 4: Precio Compra, 5: Stock Actual, 6: Stock Mínimo,
                # 7: Unidad de Medida, 8: Activo
                name         = str(row[0] or '').strip()
                description  = str(row[1] or '').strip()
                product_type = str(row[2] or 'otro').strip().lower()
                sale_price   = float(row[3] or 0)
                purchase_price = float(row[4] or 0)
                stock        = float(row[5] or 0)
                min_stock    = float(row[6] or 0)
                unit         = str(row[7] or 'unidad').strip().lower()
                is_active    = str(row[8] or 'true').strip().lower() in ['true', '1', 'si', 'sí', 'yes', 'v']

                if not name:
                    errors.append(f"Fila {i}: Falta nombre de producto")
                    continue

                # Normalizar valores a los choices válidos del modelo
                if product_type not in VALID_TYPES:
                    product_type = 'otro'
                if unit not in VALID_UNITS:
                    unit = 'unidad'

                defaults_data = {
                    'description':   description,
                    'product_type':  product_type,
                    'sale_price':    sale_price,
                    'purchase_price': purchase_price,
                    'stock':         stock,
                    'min_stock':     min_stock,
                    'unit':          unit,
                    'is_active':     is_active,
                }

                prod, c = Product.objects.get_or_create(
                    name=name,
                    defaults={'name': name, **defaults_data},
                )
                if c:
                    created += 1
                else:
                    for field, value in defaults_data.items():
                        setattr(prod, field, value)
                    prod.save()
                    updated += 1
            except Exception as e:
                errors.append(f"Fila {i}: Error procesando - {str(e)}")

        return Response({'created': created, 'updated': updated, 'errors': errors})

