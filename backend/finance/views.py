from rest_framework import generics
from rest_framework.views import APIView
from rest_framework.response import Response
from django.http import HttpResponse
from django.db.models import Sum
import openpyxl, datetime
import openpyxl, datetime, os, tempfile
from django.core.management import call_command
from .models import Transaction, CompanySettings
from .serializers import TransactionSerializer, CompanySettingsSerializer
from core.permissions import IsAdminOrVendedor
from orders.models import Order


class TransactionListCreateView(generics.ListCreateAPIView):
    serializer_class = TransactionSerializer
    permission_classes = [IsAdminOrVendedor]
    pagination_class = None

    def get_queryset(self):
        qs = Transaction.objects.all().order_by('-date')
        t = self.request.query_params.get('type')
        if t:
            qs = qs.filter(transaction_type=t)
        return qs

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)


class TransactionDetailView(generics.RetrieveUpdateDestroyAPIView):
    serializer_class = TransactionSerializer
    queryset = Transaction.objects.all()
    permission_classes = [IsAdminOrVendedor]


class FinanceSummaryView(APIView):
    permission_classes = [IsAdminOrVendedor]

    def get(self, request):
        period = request.query_params.get('period', 'month')
        today = datetime.date.today()
        if period == 'day':
            start = today
        elif period == 'week':
            start = today - datetime.timedelta(days=7)
        else:
            start = today.replace(day=1)
        qs_transactions = Transaction.objects.filter(date__gte=start)
        ingresos_manuales = float(qs_transactions.filter(transaction_type='ingreso').exclude(category='venta').aggregate(t=Sum('amount'))['t'] or 0)
        egresos = float(qs_transactions.filter(transaction_type='egreso').aggregate(t=Sum('amount'))['t'] or 0)

        payment_filter = request.query_params.get('payment_status', 'pagado')
        qs_orders = Order.objects.filter(created_at__date__gte=start)
        if payment_filter == 'pagado':
            qs_orders = qs_orders.filter(payment_status='pagado')
        ingresos_ventas = float(qs_orders.aggregate(t=Sum('total'))['t'] or 0)
        
        ingresos = ingresos_manuales + ingresos_ventas

        return Response({
            'ingresos': ingresos,
            'egresos': egresos,
            'balance': ingresos - egresos,
            'period': period,
        })

class FinanceSalesView(APIView):
    permission_classes = [IsAdminOrVendedor]

    def get(self, request):
        from orders.models import OrderItem
        payment_filter = request.query_params.get('payment_status', 'pagado')
        
        qs = OrderItem.objects.select_related('order', 'order__customer', 'product').order_by('-order__created_at')
        if payment_filter == 'pagado':
            qs = qs.filter(order__payment_status='pagado')
        
        product_filter = request.query_params.get('product_id')
        if product_filter:
            qs = qs.filter(product_id=product_filter)

        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')
        if start_date:
            qs = qs.filter(order__created_at__date__gte=start_date)
        if end_date:
            qs = qs.filter(order__created_at__date__lte=end_date)
            
        data = []
        for item in qs:
            data.append({
                'id': item.id,
                'order_id': item.order.id,
                'product_name': item.product.name,
                'quantity': float(item.quantity),
                'subtotal': float(item.subtotal),
                'margin': float(item.margin),
                'customer_name': item.order.customer.get_full_name() or item.order.customer.username,
                'payment_method': item.order.payment_method,
                'date': str(item.order.created_at.date())
            })
        return Response(data)


class ExportTransactionsView(APIView):
    permission_classes = [IsAdminOrVendedor]

    def get(self, request):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Transacciones"
        ws.append(['ID', 'Tipo', 'Categoría', 'Monto', 'Descripción', 'Referencia', 'Fecha'])
        for t in Transaction.objects.all().order_by('-date'):
            ws.append([t.id, t.transaction_type, t.category, float(t.amount),
                       t.description, t.reference_id, str(t.date)])
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="finanzas.xlsx"'
        wb.save(response)
        return response


class FinanceStatsView(APIView):
    permission_classes = [IsAdminOrVendedor]

    def get(self, request):
        from orders.models import OrderItem
        period = request.query_params.get('period', 'month')
        payment_filter = request.query_params.get('payment_status', 'pagado')
        
        qs = OrderItem.objects.all()
        if payment_filter == 'pagado':
            qs = qs.filter(order__payment_status='pagado')
        
        if period != 'all':
            today = datetime.date.today()
            if period == 'day':
                start = today
            elif period == 'week':
                start = today - datetime.timedelta(days=7)
            elif period == 'year':
                start = today.replace(month=1, day=1)
            else: # month
                start = today.replace(day=1)
            qs = qs.filter(order__created_at__date__gte=start)
            
        stats = list(
            qs.values('product__name')
            .annotate(
                total_quantity=Sum('quantity'),
                total_revenue=Sum('subtotal'),
                total_profit=Sum('margin')
            )
            .order_by('-total_profit')
        )
        
        # Calculate summary metrics
        total_revenue = sum(float(item['total_revenue'] or 0) for item in stats)
        total_profit = sum(float(item['total_profit'] or 0) for item in stats)
        total_quantity = sum(float(item['total_quantity'] or 0) for item in stats)
        
        return Response({
            'period': period,
            'summary': {
                'total_revenue': total_revenue,
                'total_profit': total_profit,
                'total_quantity': total_quantity,
                'avg_margin_pct': (total_profit / total_revenue * 100) if total_revenue > 0 else 0
            },
            'product_stats': stats
        })

class CompanySettingsView(APIView):
    permission_classes = [IsAdminOrVendedor]

    def get(self, request):
        settings_obj = CompanySettings.load()
        serializer = CompanySettingsSerializer(settings_obj)
        return Response(serializer.data)

    def put(self, request):
        settings_obj = CompanySettings.load()
        serializer = CompanySettingsSerializer(settings_obj, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=400)


class DatabaseBackupView(APIView):
    permission_classes = [IsAdminOrVendedor]

    def get(self, request):
        import io
        out = io.StringIO()
        call_command('dumpdata', exclude=['contenttypes', 'auth.Permission', 'sessions'], format='json', indent=2, stdout=out)
        response = HttpResponse(out.getvalue(), content_type='application/json')
        response['Content-Disposition'] = 'attachment; filename="backup.json"'
        return response

    def post(self, request):
        if 'file' not in request.FILES:
            return Response({'error': 'No file provided'}, status=400)
            
        file = request.FILES['file']
        fd, path = tempfile.mkstemp(suffix='.json')
        try:
            with os.fdopen(fd, 'wb') as tmp:
                for chunk in file.chunks():
                    tmp.write(chunk)
            call_command('loaddata', path)
            return Response({'status': 'Database restored successfully'})
        except Exception as e:
            return Response({'error': str(e)}, status=500)
        finally:
            os.remove(path)


class DownloadFinanceTemplateView(APIView):
    permission_classes = [IsAdminOrVendedor]

    def get(self, request):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Plantilla Finanzas"
        # IMPORTANTE: el orden de estas columnas debe coincidir exactamente
        # con lo que lee ImportFinanceView (min_row=2, por índice de columna).
        headers = [
            'Tipo',           # col 0 — ingreso | egreso
            'Categoría',      # col 1 — venta | compra | gasto_operacional | combustible | cajas | despacho | marketing | otro
            'Monto',          # col 2 — número positivo
            'Descripción',    # col 3 — texto libre
            'Referencia',     # col 4 — opcional (ID de pedido, factura, etc.)
            'Fecha (YYYY-MM-DD)',  # col 5
        ]
        ws.append(headers)

        # Filas de ejemplo para guiar al usuario
        today = datetime.date.today().isoformat()
        ws.append(['ingreso', 'venta',              15000, 'Venta de paltas Hass',          'PED-001', today])
        ws.append(['egreso',  'gasto_operacional',   3500, 'Pago de luz del mes',            '',        today])
        ws.append(['egreso',  'combustible',         8000, 'Gasolina camioneta delivery',    '',        today])
        ws.append(['egreso',  'cajas',               2000, 'Compra de cajas para embalaje',  '',        today])
        ws.append(['egreso',  'despacho',            5000, 'Servicio de courier externo',    '',        today])
        ws.append(['egreso',  'marketing',           4000, 'Publicidad en redes sociales',   '',        today])
        ws.append(['ingreso', 'otro',                1000, 'Ingreso misceláneo',             '',        today])

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="plantilla_finanzas.xlsx"'
        wb.save(response)
        return response


class ImportFinanceView(APIView):
    permission_classes = [IsAdminOrVendedor]

    VALID_TYPES = {'ingreso', 'egreso'}
    VALID_CATEGORIES = {
        'venta', 'compra', 'gasto_operacional', 'combustible',
        'cajas', 'despacho', 'marketing', 'otro'
    }

    def post(self, request):
        file = request.FILES.get('file')
        import_mode = request.data.get('import_mode', 'update')

        if not file:
            return Response({'error': 'No se recibió archivo'}, status=400)

        if import_mode == 'replace':
            # Solo eliminar transacciones manuales; las de ventas se preservan
            Transaction.objects.exclude(category='venta').delete()

        wb = openpyxl.load_workbook(file)
        ws = wb.active
        created = 0
        errors = []

        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                # Columnas según la plantilla (mismo orden):
                # 0: Tipo, 1: Categoría, 2: Monto, 3: Descripción, 4: Referencia, 5: Fecha
                transaction_type = str(row[0] or '').strip().lower()
                category         = str(row[1] or '').strip().lower()
                amount           = float(row[2] or 0)
                description      = str(row[3] or '').strip()
                reference_id     = str(row[4] or '').strip()
                date_str         = str(row[5] or '').strip()

                if not transaction_type and not description and not amount:
                    continue  # Fila vacía (e.g. fila de ejemplo)

                # Validaciones
                if not description:
                    errors.append(f"Fila {i}: Falta descripción")
                    continue
                if amount <= 0:
                    errors.append(f"Fila {i}: El monto debe ser mayor a 0")
                    continue
                if transaction_type not in self.VALID_TYPES:
                    errors.append(f"Fila {i}: Tipo '{transaction_type}' inválido (debe ser 'ingreso' o 'egreso')")
                    continue
                if category not in self.VALID_CATEGORIES:
                    category = 'otro'  # Normalizar silenciosamente

                # Parsear fecha
                try:
                    tx_date = datetime.datetime.strptime(date_str, '%Y-%m-%d').date()
                except (ValueError, TypeError):
                    tx_date = datetime.date.today()

                Transaction.objects.create(
                    transaction_type=transaction_type,
                    category=category,
                    amount=amount,
                    description=description,
                    reference_id=reference_id,
                    date=tx_date,
                    created_by=request.user,
                )
                created += 1
            except Exception as e:
                errors.append(f"Fila {i}: Error procesando - {str(e)}")

        return Response({'created': created, 'errors': errors})


class DownloadSalesTemplateView(APIView):
    permission_classes = [IsAdminOrVendedor]

    def get(self, request):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Plantilla Ventas Detalladas"
        
        headers = [
            'ID Grupo Pedido (Opcional)',  # col 0
            'Fecha (YYYY-MM-DD)',          # col 1
            'Cliente (Email o Teléfono)',  # col 2
            'Producto (Nombre)',           # col 3
            'Cantidad',                    # col 4
            'Subtotal (Opcional)',         # col 5
            'Medio Pago',                  # col 6 (efectivo, transferencia, transbank)
            'Estado Pago'                  # col 7 (pagado, abonado, pendiente)
        ]
        ws.append(headers)

        today = datetime.date.today().isoformat()
        ws.append(['PED-001', today, 'cliente@gmail.com', 'Palta Hass', 2, 10000, 'efectivo', 'pagado'])
        ws.append(['PED-001', today, 'cliente@gmail.com', 'Malla Limón', 1, 3000, 'efectivo', 'pagado'])
        ws.append(['', today, '+56912345678', 'Huevo Extra', 1, 5000, 'transferencia', 'pagado'])

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="plantilla_ventas.xlsx"'
        wb.save(response)
        return response


class ImportSalesView(APIView):
    permission_classes = [IsAdminOrVendedor]

    def post(self, request):
        file = request.FILES.get('file')
        import_mode = request.data.get('import_mode', 'update')

        if not file:
            return Response({'error': 'No se recibió archivo'}, status=400)

        from orders.models import Order, OrderItem
        from products.models import Product
        from django.contrib.auth import get_user_model
        from django.db.models import Q
        import decimal

        User = get_user_model()

        if import_mode == 'replace':
            # Limpiar pedidos, order items, y transacciones de venta
            Transaction.objects.filter(category='venta').delete()
            Order.objects.all().delete()
            # Note: OrderItem deletes on cascade with Order

        wb = openpyxl.load_workbook(file)
        ws = wb.active
        created_items = 0
        errors = []

        # Cache para agrupar pedidos de la misma fila (si comparten ID)
        orders_cache = {}

        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                group_id       = str(row[0] or '').strip()
                date_str       = str(row[1] or '').strip()
                customer_id    = str(row[2] or '').strip()
                product_name   = str(row[3] or '').strip()
                quantity       = float(row[4] or 0)
                subtotal_raw   = row[5]
                payment_method = str(row[6] or 'efectivo').strip().lower()
                payment_status = str(row[7] or 'pagado').strip().lower()

                if not customer_id and not product_name and not quantity:
                    continue  # fila vacía

                if not customer_id:
                    errors.append(f"Fila {i}: Falta cliente")
                    continue
                if not product_name:
                    errors.append(f"Fila {i}: Falta producto")
                    continue
                if quantity <= 0:
                    errors.append(f"Fila {i}: Cantidad debe ser > 0")
                    continue

                # Buscar Cliente
                digits = ''.join(filter(str.isdigit, customer_id))
                base_phone = digits[-9:] if len(digits) >= 9 else digits
                query = Q(email__iexact=customer_id)
                if base_phone:
                    query |= Q(phone__endswith=base_phone) | Q(whatsapp_number__endswith=base_phone)
                
                customer = User.objects.filter(query).first()
                if not customer:
                    errors.append(f"Fila {i}: Cliente no encontrado ({customer_id})")
                    continue

                # Buscar Producto
                product = Product.objects.filter(name__iexact=product_name).first()
                if not product:
                    errors.append(f"Fila {i}: Producto no encontrado ({product_name})")
                    continue

                # Calcular fechas y montos
                try:
                    order_date = datetime.datetime.strptime(date_str, '%Y-%m-%d').date()
                    order_dt = datetime.datetime.strptime(date_str, '%Y-%m-%d')
                except (ValueError, TypeError):
                    order_date = datetime.date.today()
                    order_dt = datetime.datetime.now()

                subtotal = float(subtotal_raw) if subtotal_raw else float(product.sale_price) * quantity

                # Calcular costo unitario (igual que al crear orden)
                if product.is_bundle:
                    bundle_cost = sum([float(c.product.purchase_price or 0) * float(c.quantity) for c in product.components.all()])
                    unit_cost = bundle_cost
                else:
                    unit_cost = float(product.purchase_price) if product.purchase_price else 0

                # Agrupar en Order o crear uno nuevo
                order_key = group_id if group_id else f"ROW_{i}"
                if order_key not in orders_cache:
                    order = Order.objects.create(
                        customer=customer,
                        total=0,
                        subtotal=0,
                        delivery_cost=0,
                        status='entregado',
                        payment_method=payment_method,
                        payment_status=payment_status,
                    )
                    order.created_at = order_dt
                    order.save(update_fields=['created_at'])
                    orders_cache[order_key] = order
                else:
                    order = orders_cache[order_key]

                # Crear OrderItem
                oi = OrderItem.objects.create(
                    order=order,
                    product=product,
                    quantity=quantity,
                    unit_price=subtotal / quantity if quantity > 0 else float(product.sale_price),
                    unit_cost=unit_cost,
                )

                # Actualizar el total del pedido
                order.subtotal = float(order.subtotal) + subtotal
                order.total = float(order.total) + subtotal
                order.save(update_fields=['subtotal', 'total'])

                # Descontar stock
                if product.is_bundle:
                    for comp in product.components.all():
                        comp.product.stock -= decimal.Decimal(str(quantity * float(comp.quantity)))
                        comp.product.save()
                else:
                    product.stock -= decimal.Decimal(str(quantity))
                    product.save()

                created_items += 1
            except Exception as e:
                errors.append(f"Fila {i}: Error - {str(e)}")

        # Ahora que todos los items están agregados a las Orders, generamos las Transactions
        for order in orders_cache.values():
            if order.payment_status in ('pagado', 'abonado') and order.total > 0:
                if not Transaction.objects.filter(reference_id=str(order.id), category='venta').exists():
                    Transaction.objects.create(
                        transaction_type='ingreso',
                        category='venta',
                        amount=order.total,
                        description=f'Pedido #{order.id} (importado)',
                        reference_id=str(order.id),
                        date=order.created_at.date(),
                        created_by=request.user,
                    )

        return Response({'created': created_items, 'errors': errors})

