from rest_framework import generics, status
from rest_framework.views import APIView
from rest_framework.response import Response
from django.http import HttpResponse
from django.conf import settings
import mercadopago
import openpyxl
from .models import Order, OrderItem
from .serializers import OrderSerializer, OrderItemSerializer
from products.models import Product
from finance.models import Transaction
from rest_framework.permissions import AllowAny, IsAuthenticated
from core.permissions import IsAdminOrVendedor, IsOwnerOrAdmin
from core.permissions import IsAdminOrVendedor, IsOwnerOrAdmin
import datetime
from rest_framework import filters


class OrderListCreateView(generics.ListCreateAPIView):
    """
    Clientes solo ven sus pedidos; admin/vendedor ven todos.
    La lógica de filtrado ya existe en get_queryset — solo necesitamos
    que sea IsAuthenticated para que cualquier usuario logueado pueda crear
    su propio pedido.
    """
    serializer_class = OrderSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['id', 'customer__first_name', 'customer__last_name', 'customer__username', 'customer__email', 'customer__phone']
    ordering_fields = ['id', 'created_at', 'total', 'status', 'payment_status', 'customer__first_name']
    ordering = ['-created_at']

    def get_queryset(self):
        user = self.request.user
        if user.role in ['admin', 'vendedor']:
            qs = Order.objects.all()
            status_filter = self.request.query_params.get('status')
            if status_filter:
                qs = qs.filter(status=status_filter)
            return qs
        return Order.objects.filter(customer=user)

    def create(self, request, *args, **kwargs):
        data = request.data
        items_data = data.get('items', [])
        
        # Security: Prevent customer ID spoofing
        if request.user.role in ['admin', 'vendedor']:
            customer_id = data.get('customer_id', request.user.id)
        else:
            customer_id = request.user.id
            
        # Pre-validate stock
        from decimal import Decimal
        for item in items_data:
            try:
                product = Product.objects.get(id=item['product_id'])
                qty = Decimal(str(item['quantity']))
                if product.is_bundle:
                    for comp in product.components.all():
                        required = qty * comp.quantity
                        if comp.product.stock < required:
                            return Response({'error': f'Stock insuficiente de {comp.product.name} para el combo {product.name}'}, status=status.HTTP_400_BAD_REQUEST)
                else:
                    if product.stock < qty:
                        return Response({'error': f'Stock insuficiente para {product.name}'}, status=status.HTTP_400_BAD_REQUEST)
            except Product.DoesNotExist:
                return Response({'error': 'Producto no encontrado'}, status=status.HTTP_400_BAD_REQUEST)

        order = Order.objects.create(
            customer_id=customer_id,
            delivery_type=data.get('delivery_type', 'retiro'),
            delivery_address=data.get('delivery_address', ''),
            delivery_commune=data.get('delivery_commune', ''),
            delivery_reference=data.get('delivery_reference', ''),
            payment_method=data.get('payment_method', 'efectivo'),
            payment_condition=data.get('payment_condition', 'inmediato'),
            notes=data.get('notes', ''),
        )
        subtotal = 0
        for item in items_data:
            product = Product.objects.get(id=item['product_id'])
            qty = Decimal(str(item['quantity']))
            # Security: Always use product.sale_price instead of user input
            unit_price = Decimal(str(product.sale_price))
            # A5 fix: usar order_by explícito para obtener el costo más reciente
            if product.is_bundle:
                bundle_cost = Decimal('0')
                for comp in product.components.all():
                    comp_cost = comp.product.purchase_price if comp.product.purchase_price else Decimal('0')
                    bundle_cost += comp_cost * comp.quantity
                    
                    comp.product.stock -= (qty * comp.quantity)
                    comp.product.save()
                unit_cost = bundle_cost
            else:
                unit_cost = product.purchase_price if product.purchase_price else Decimal('0')
                
                product.stock -= qty
                product.save()
                
            oi = OrderItem.objects.create(
                order=order, product=product, quantity=qty,
                unit_price=unit_price, unit_cost=unit_cost,
            )
            subtotal += float(oi.subtotal)
        order.subtotal = subtotal
        order.total = subtotal + float(order.delivery_cost)
        order.save()
        Transaction.objects.create(
            transaction_type='ingreso', category='venta',
            amount=order.total, description=f'Pedido #{order.id}',
            reference_id=str(order.id), date=datetime.date.today(),
            created_by=request.user,
        )
        return Response(OrderSerializer(order).data, status=status.HTTP_201_CREATED)


class OrderDetailView(generics.RetrieveUpdateDestroyAPIView):
    """
    El propietario del pedido o un admin/vendedor pueden ver/editar/eliminar.
    """
    serializer_class = OrderSerializer
    queryset = Order.objects.all()
    permission_classes = [IsOwnerOrAdmin]

    def perform_update(self, serializer):
        # Security: Prevent non-staff users from updating restricted fields
        user = self.request.user
        if user.role not in ['admin', 'vendedor']:
            restricted_fields = ['status', 'payment_status', 'total', 'subtotal', 'delivery_cost', 'points_earned', 'mercadopago_preference_id', 'mercadopago_payment_id', 'mercadopago_link']
            for field in restricted_fields:
                if field in serializer.validated_data:
                    serializer.validated_data.pop(field)
        instance = serializer.save()
        if instance.status == 'entregado' and instance.payment_status == 'pagado' and not instance.points_awarded:
            from loyalty.models import LoyaltyAccount, PointTransaction
            points = int(instance.total / 1000) * 10
            if points > 0:
                instance.points_earned = points
                instance.points_awarded = True
                instance.save(update_fields=['points_earned', 'points_awarded'])
                acc, _ = LoyaltyAccount.objects.get_or_create(user=instance.customer)
                acc.points += points
                acc.total_points_earned += points
                acc.total_purchases += instance.total
                acc.update_level()
                PointTransaction.objects.create(
                    account=acc,
                    transaction_type='ganado',
                    points=points,
                    description=f'Pedido #{instance.id} Entregado',
                    reference_id=str(instance.id)
                )

    def perform_destroy(self, instance):
        # Devolver el stock
        from decimal import Decimal
        for item in instance.items.all():
            product = item.product
            if product.is_bundle:
                for comp in product.components.all():
                    comp.product.stock += Decimal(str(item.quantity * comp.quantity))
                    comp.product.save()
            else:
                product.stock += Decimal(str(item.quantity))
                product.save()
            
        # Eliminar transacciones financieras asociadas
        try:
            from finance.models import Transaction
            Transaction.objects.filter(reference_id=f"ORDER-{instance.id}").delete()
        except Exception as e:
            pass
        
        # Opcional: Eliminar puntos obtenidos si aplicara
        try:
            if instance.points_earned > 0:
                acc = instance.customer.loyalty
                acc.points -= instance.points_earned
                acc.total_points_earned -= instance.points_earned
                acc.total_purchases -= instance.total
                acc.update_level()
        except Exception:
            pass

        instance.delete()


class GenerateMercadoPagoView(APIView):
    """Solo el propietario o staff pueden generar el link de pago."""
    permission_classes = [IsOwnerOrAdmin]

    def post(self, request, pk):
        order = Order.objects.get(pk=pk)
        self.check_object_permissions(request, order)
        sdk = mercadopago.SDK(settings.MERCADOPAGO_ACCESS_TOKEN)
        preference_data = {
            "items": [{"title": f"Pedido #{order.id} - Palta con Huevo", "quantity": 1, "unit_price": float(order.total)}],
            "payer": {"email": order.customer.email},
            "back_urls": {
                "success": f"{settings.WHATSAPP_SERVICE_URL}/mp/success",
                "failure": f"{settings.WHATSAPP_SERVICE_URL}/mp/failure",
            },
            "auto_return": "approved",
            "notification_url": f"{settings.WHATSAPP_SERVICE_URL}/api/orders/webhook/mercadopago/",
            "external_reference": str(order.id),
        }
        result = sdk.preference().create(preference_data)
        preference = result["response"]
        order.mercadopago_preference_id = preference["id"]
        order.mercadopago_link = preference["init_point"]
        order.save()
        return Response({"link": preference["init_point"], "preference_id": preference["id"]})


from loyalty.models import LoyaltyAccount, PointTransaction


class MercadoPagoWebhookView(APIView):
    """
    Webhook externo de MercadoPago — debe ser AllowAny.
    Procesa notificaciones de pago aprobadas de forma idempotente y acredita los puntos de lealtad al cliente.
    """
    permission_classes = [AllowAny]

    def post(self, request):
        data = request.data
        if data.get('type') == 'payment':
            payment_id = data.get('data', {}).get('id')
            if not payment_id:
                return Response({'status': 'ignored'}, status=status.HTTP_200_OK)

            try:
                sdk = mercadopago.SDK(settings.MERCADOPAGO_ACCESS_TOKEN)
                payment_resp = sdk.payment().get(payment_id)
                payment = payment_resp.get("response", {})
            except Exception:
                # Si falla la comunicación con la API de MercadoPago en testing/mock
                payment = {}

            if payment.get('status') == 'approved':
                order_id = payment.get('external_reference')
                if order_id:
                    try:
                        order = Order.objects.get(id=order_id)
                        # Check idempotencia: Si ya está pagado, no volver a procesar
                        if order.payment_status == 'pagado':
                            return Response({'status': 'already_processed'}, status=status.HTTP_200_OK)

                        order.payment_status = 'pagado'
                        order.mercadopago_payment_id = str(payment_id)
                        order.save()

                        # Acreditar puntos de lealtad se movió a perform_update cuando el pedido se marque como entregado
                        # Aquí solo marcamos como pagado.
                    except Order.DoesNotExist:
                        pass
        return Response({'status': 'ok'})


class ExportOrdersView(APIView):
    permission_classes = [IsAdminOrVendedor]

    def get(self, request):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Pedidos"
        # IMPORTANTE: el orden de estas columnas coincide exactamente con
        # lo que lee ImportOrdersView (min_row=2, por índice de columna).
        ws.append([
            'Email o Teléfono Cliente',  # col 0 — identificador del cliente
            'Total',                     # col 1
            'Costo Envío',               # col 2
            'Descuento',                 # col 3
            'Estado',                    # col 4 — pendiente|entregado|cancelado
            'Método Pago',               # col 5 — efectivo|transferencia|mercadopago
            'Estado Pago',               # col 6 — pendiente|abonado|pagado|vencido
            'Fecha Creación (YYYY-MM-DD)',  # col 7
        ])
        for o in Order.objects.all().order_by('-created_at'):
            # Usar email como identificador principal; caer a teléfono si no tiene email
            identifier = o.customer.email or o.customer.phone or o.customer.username
            ws.append([
                identifier,
                float(o.total),
                float(o.delivery_cost),
                float(o.discount),
                o.status,
                o.payment_method,
                o.payment_status,
                str(o.created_at.date()),
            ])
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="pedidos.xlsx"'
        wb.save(response)
        return response


class DashboardView(APIView):
    permission_classes = [IsAdminOrVendedor]

    def get(self, request):
        from django.db.models import Sum, Count
        from django.utils import timezone
        import datetime
        from products.models import Product
        from users.models import User
        from .serializers import OrderSerializer
        from .models import Order, OrderItem

        today = datetime.date.today()
        month_start = today.replace(day=1)
        
        payment_filter = request.query_params.get('payment_status', 'pagado')
        
        orders_today = Order.objects.filter(created_at__date=today)
        orders_month = Order.objects.filter(created_at__date__gte=month_start)
        
        # Filtro de ventas dinámico para la vista aislada (por defecto Mes)
        sales_period = request.query_params.get('sales_period', 'month')
        if sales_period == 'day':
            sales_start_date = today
        elif sales_period == 'week':
            sales_start_date = today - datetime.timedelta(days=7)
        else: # month
            sales_start_date = month_start
            
        sales_period_orders = Order.objects.filter(created_at__date__gte=sales_start_date)
        
        if payment_filter == 'pagado':
            orders_today = orders_today.filter(payment_status='pagado')
            orders_month = orders_month.filter(payment_status='pagado')
            sales_period_orders = sales_period_orders.filter(payment_status='pagado')
            
        sales_period_value = float(sales_period_orders.aggregate(t=Sum('total'))['t'] or 0)
        
        # Pedidos pagados en el período (suma de totales)
        sales_paid = float(Order.objects.filter(created_at__date__gte=sales_start_date, payment_status='pagado').aggregate(t=Sum('total'))['t'] or 0)
        
        # Pedidos entregados pero no pagados en el período
        sales_unpaid = float(Order.objects.filter(created_at__date__gte=sales_start_date, status='entregado').exclude(payment_status='pagado').aggregate(t=Sum('total'))['t'] or 0)
        
        # Pedidos pendientes (no entregados y no pagados) (cantidad)
        pending_qs = Order.objects.exclude(status='entregado').exclude(payment_status='pagado')
        orders_pending = pending_qs.count()
        orders_pending_value = float(pending_qs.aggregate(t=Sum('total'))['t'] or 0)

        # Listado de pedidos pendientes: no entregados, ordenados por más recientes
        pending_orders = Order.objects.exclude(status__in=['entregado', 'cancelado']).order_by('-created_at')[:10]

        products_sold_qs = OrderItem.objects.filter(order__created_at__date__gte=sales_start_date)
        if payment_filter == 'pagado':
            products_sold_qs = products_sold_qs.filter(order__payment_status='pagado')

        products_sold = list(
            products_sold_qs
            .values('product__name')
            .annotate(
                total_quantity=Sum('quantity'),
                total_sales=Sum('subtotal')
            )
            .order_by('-total_sales')
        )

        top_customers_qs = Order.objects.all()
        if payment_filter == 'pagado':
            top_customers_qs = top_customers_qs.filter(payment_status='pagado')

        top_customers = list(
            top_customers_qs
            .values('customer__first_name', 'customer__last_name', 'customer__username')
            .annotate(total_spent=Sum('total'), total_orders=Count('id'))
            .order_by('-total_spent')[:5]
        )

        return Response({
            'sales_period_value': sales_period_value,
            'sales_period': sales_period,
            'sales_today': float(orders_today.aggregate(t=Sum('total'))['t'] or 0),
            'sales_month': float(orders_month.aggregate(t=Sum('total'))['t'] or 0),
            'orders_today': orders_today.count(),
            'sales_paid': sales_paid,
            'sales_unpaid': sales_unpaid,
            'orders_pending': orders_pending,
            'orders_pending_value': orders_pending_value,
            'total_customers': User.objects.filter(role='cliente').count(),
            'low_stock_count': sum(1 for p in Product.objects.filter(is_active=True) if p.stock <= p.min_stock),
            'pending_delivery_orders': OrderSerializer(pending_orders, many=True).data,
            'products_sold': products_sold,
            'top_customers': top_customers,
        })

class OrderPaymentCreateView(APIView):
    permission_classes = [IsAdminOrVendedor]
    def post(self, request, pk):
        from django.shortcuts import get_object_or_404
        order = get_object_or_404(Order, pk=pk)
        amount = request.data.get('amount')
        payment_method = request.data.get('payment_method')
        notes = request.data.get('notes', '')
        if not amount or not payment_method:
            return Response({'error': 'amount y payment_method requeridos'}, status=status.HTTP_400_BAD_REQUEST)
        
        from .models import OrderPayment
        payment = OrderPayment.objects.create(
            order=order,
            amount=amount,
            payment_method=payment_method,
            notes=notes,
            created_by=request.user
        )
        from .serializers import OrderPaymentSerializer
        return Response(OrderPaymentSerializer(payment).data)

class OrderItemUpdateView(APIView):
    permission_classes = [IsAdminOrVendedor]
    def patch(self, request, pk, item_id):
        from django.shortcuts import get_object_or_404
        order = get_object_or_404(Order, pk=pk)
        from .models import OrderItem
        item = get_object_or_404(OrderItem, pk=item_id, order=order)
        status_val = request.data.get('status')
        quantity = request.data.get('quantity')
        unit_price = request.data.get('unit_price')
        
        needs_order_recalc = False
        
        if status_val:
            item.status = status_val
        if quantity is not None:
            item.quantity = quantity
            needs_order_recalc = True
        if unit_price is not None:
            item.unit_price = unit_price
            needs_order_recalc = True
            
        item.save()
        
        if needs_order_recalc:
            order.subtotal = sum(i.subtotal for i in order.items.all())
            order.total = order.subtotal + order.delivery_cost - order.discount
            order.save()
        
        from .serializers import OrderItemSerializer
        return Response(OrderItemSerializer(item).data)

class DownloadOrderTemplateView(APIView):
    permission_classes = [IsAdminOrVendedor]

    def get(self, request):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Plantilla Pedidos"
        # IMPORTANTE: el orden de estas columnas debe coincidir exactamente
        # con lo que lee ImportOrdersView (min_row=2, por índice de columna).
        headers = [
            'Email o Teléfono Cliente',      # col 0 — obligatorio
            'Total',                          # col 1
            'Costo Envío',                    # col 2
            'Descuento',                      # col 3
            'Estado',                         # col 4 — pendiente|entregado|cancelado
            'Método Pago',                    # col 5 — efectivo|transferencia|mercadopago
            'Estado Pago',                    # col 6 — pendiente|abonado|pagado|vencido
            'Fecha Creación (YYYY-MM-DD)',     # col 7
        ]
        ws.append(headers)

        # Fila de ejemplo para guiar al usuario
        ws.append([
            'cliente@ejemplo.com', 15000, 0, 0,
            'entregado', 'efectivo', 'pagado', '2025-01-15'
        ])

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="plantilla_pedidos.xlsx"'
        wb.save(response)
        return response



class ImportOrdersView(APIView):
    permission_classes = [IsAdminOrVendedor]

    def post(self, request):
        file = request.FILES.get('file')
        import_mode = request.data.get('import_mode', 'update')

        if not file:
            return Response({'error': 'No se recibió archivo'}, status=400)

        if import_mode == 'replace':
            # Al reemplazar, limpiar también las transacciones de ventas asociadas a pedidos
            from finance.models import Transaction
            Transaction.objects.filter(category='venta').delete()
            Order.objects.all().delete()

        wb = openpyxl.load_workbook(file)
        ws = wb.active
        created = 0
        errors = []

        from django.contrib.auth import get_user_model
        from django.db.models import Q
        from finance.models import Transaction
        import datetime as dt_module

        User = get_user_model()

        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                identifier    = str(row[0] or '').strip()
                total         = float(row[1] or 0)
                delivery_cost = float(row[2] or 0)
                discount      = float(row[3] or 0)
                status_val    = str(row[4] or 'entregado').strip().lower()
                payment_method = str(row[5] or 'efectivo').strip().lower()
                payment_status = str(row[6] or 'pagado').strip().lower()
                created_at_str = str(row[7] or '').strip()

                if not identifier:
                    errors.append(f"Fila {i}: Falta identificador del cliente")
                    continue

                # Buscar cliente por email o teléfono
                digits = ''.join(filter(str.isdigit, identifier))
                base_phone = digits[-9:] if len(digits) >= 9 else digits

                query = Q(email__iexact=identifier)
                if base_phone:
                    query |= Q(phone__endswith=base_phone) | Q(whatsapp_number__endswith=base_phone)

                customer = User.objects.filter(query).first()

                if not customer:
                    errors.append(f"Fila {i}: Cliente no encontrado ({identifier})")
                    continue

                # Determinar la fecha del pedido
                order_date = dt_module.date.today()
                if created_at_str:
                    try:
                        order_date = dt_module.datetime.strptime(created_at_str, '%Y-%m-%d').date()
                    except Exception:
                        pass

                order = Order.objects.create(
                    customer=customer,
                    total=total,
                    subtotal=total - delivery_cost + discount,
                    delivery_cost=delivery_cost,
                    discount=discount,
                    status=status_val,
                    payment_method=payment_method,
                    payment_status=payment_status,
                )

                # Ajustar la fecha de creación al valor del archivo
                if created_at_str:
                    try:
                        order_dt = dt_module.datetime.strptime(created_at_str, '%Y-%m-%d')
                        order.created_at = order_dt
                        order.save(update_fields=['created_at'])
                    except Exception:
                        pass

                # ── SINCRONIZACIÓN CON FINANZAS ──────────────────────────────
                # Crear Transaction de venta para que el dashboard, finanzas y
                # estadísticas incluyan este pedido histórico.
                # Solo se registra si hubo algún pago (pagado o abonado).
                if payment_status in ('pagado', 'abonado') and total > 0:
                    # Evitar duplicados: no crear si ya existe para este pedido
                    if not Transaction.objects.filter(reference_id=str(order.id), category='venta').exists():
                        Transaction.objects.create(
                            transaction_type='ingreso',
                            category='venta',
                            amount=total,
                            description=f'Pedido #{order.id} (importado)',
                            reference_id=str(order.id),
                            date=order_date,
                            created_by=request.user,
                        )

                created += 1
            except Exception as e:
                errors.append(f"Fila {i}: Error procesando - {str(e)}")

        return Response({'created': created, 'errors': errors})

